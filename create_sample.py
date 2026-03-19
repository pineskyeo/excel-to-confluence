"""
Sample Excel Creator
테이블, 이미지, 텍스트가 포함된 샘플 Excel 파일 생성
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage, ImageDraw, ImageFont


def make_image(title: str, width=400, height=180, bg: tuple = (70, 130, 180)) -> io.BytesIO:
    """Simple colored banner image with title text."""
    img = PILImage.new("RGB", (width, height), bg)
    draw = ImageDraw.Draw(img)

    # border
    draw.rectangle([4, 4, width - 5, height - 5], outline=(255, 255, 255), width=3)

    # centered text (basic, no custom font needed)
    text = title
    bbox = draw.textbbox((0, 0), text)
    tw = bbox[2] - bbox[0]
    th = bbox[3] - bbox[1]
    draw.text(((width - tw) // 2, (height - th) // 2), text, fill=(255, 255, 255))

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf


def header_style(cell, bg_hex: str = "4472C4"):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def create_sample():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Project Overview ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Project Overview"
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 16

    # Title text
    ws1["A1"] = "📋 프로젝트 개요 보고서"
    ws1["A1"].font = Font(bold=True, size=18, color="1F3864")

    ws1["A2"] = "작성일: 2026-03-19    작성자: 개발팀"
    ws1["A2"].font = Font(size=10, italic=True, color="595959")

    ws1["A3"] = "본 문서는 신규 플랫폼 개발 프로젝트의 현황을 정리한 보고서입니다."
    ws1["A3"].font = Font(size=11)

    # Embedded image #1
    img1_buf = make_image("Project Timeline Diagram", bg=(41, 128, 185))
    xl_img1 = XLImage(img1_buf)
    xl_img1.width = 400
    xl_img1.height = 180
    ws1.add_image(xl_img1, "A5")

    # Table — project tasks (starts at row 15 to sit below image)
    ws1["A15"] = "업무 목록"
    ws1["A15"].font = Font(bold=True, size=13, color="1F3864")

    headers = ["업무명", "담당자", "상태", "완료 예정일"]
    for col, h in enumerate(headers, start=1):
        c = ws1.cell(row=16, column=col, value=h)
        header_style(c, "4472C4")
        c.border = thin_border()

    rows = [
        ("요구사항 분석", "홍길동", "✅ 완료", "2026-01-15"),
        ("시스템 설계", "김철수", "✅ 완료", "2026-02-10"),
        ("백엔드 개발", "이영희", "🔄 진행중", "2026-03-30"),
        ("프론트엔드 개발", "박민준", "🔄 진행중", "2026-04-15"),
        ("통합 테스트", "최수진", "⏳ 대기", "2026-05-01"),
        ("운영 배포", "전체팀", "⏳ 대기", "2026-05-20"),
    ]
    alt_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    for i, row in enumerate(rows, start=17):
        for j, val in enumerate(row, start=1):
            c = ws1.cell(row=i, column=j, value=val)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                c.fill = alt_fill

    ws1["A24"] = "※ 상태는 매주 월요일 업데이트됩니다."
    ws1["A24"].font = Font(size=9, italic=True, color="808080")

    # ── Sheet 2: Technical Specs ───────────────────────────────────────
    ws2 = wb.create_sheet("Technical Specs")
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 22

    ws2["A1"] = "⚙️ 기술 사양서"
    ws2["A1"].font = Font(bold=True, size=18, color="1F3864")

    ws2["A2"] = "시스템 구성 요소 및 기술 스택 정보"
    ws2["A2"].font = Font(size=11, italic=True)

    # Merged section header
    ws2.merge_cells("A4:D4")
    ws2["A4"] = "기술 스택 (Tech Stack)"
    ws2["A4"].font = Font(bold=True, size=13, color="FFFFFF")
    ws2["A4"].fill = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    ws2["A4"].alignment = Alignment(horizontal="center")

    tech_headers = ["구분", "기술", "버전", "비고"]
    for col, h in enumerate(tech_headers, start=1):
        c = ws2.cell(row=5, column=col, value=h)
        header_style(c, "70AD47")
        c.border = thin_border()

    tech_data = [
        ("언어", "Python", "3.11+", "타입 힌트 사용"),
        ("Excel 파싱", "openpyxl", "3.1+", "이미지·병합셀 지원"),
        ("Word 생성", "python-docx", "1.1+", "Confluence 임포트용"),
        ("이미지 처리", "Pillow", "10.0+", "포맷 변환"),
        ("런타임", "Docker", "24.0", "컨테이너 배포"),
        ("CI/CD", "GitHub Actions", "latest", "자동 빌드"),
    ]
    for i, row in enumerate(tech_data, start=6):
        for j, val in enumerate(row, start=1):
            c = ws2.cell(row=i, column=j, value=val)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                c.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Embedded image #2
    img2_buf = make_image("System Architecture", width=400, height=180, bg=(39, 174, 96))
    xl_img2 = XLImage(img2_buf)
    xl_img2.width = 400
    xl_img2.height = 180
    ws2.add_image(xl_img2, "A13")

    # Second table — API endpoints
    ws2.merge_cells("A21:D21")
    ws2["A21"] = "API 엔드포인트 목록"
    ws2["A21"].font = Font(bold=True, size=13, color="FFFFFF")
    ws2["A21"].fill = PatternFill(start_color="2E4053", end_color="2E4053", fill_type="solid")
    ws2["A21"].alignment = Alignment(horizontal="center")

    api_headers = ["Method", "Endpoint", "설명", "인증"]
    for col, h in enumerate(api_headers, start=1):
        c = ws2.cell(row=22, column=col, value=h)
        header_style(c, "2980B9")
        c.border = thin_border()

    api_data = [
        ("GET", "/api/v1/users", "사용자 목록 조회", "Bearer Token"),
        ("POST", "/api/v1/users", "사용자 생성", "Bearer Token"),
        ("PUT", "/api/v1/users/{id}", "사용자 수정", "Bearer Token"),
        ("DELETE", "/api/v1/users/{id}", "사용자 삭제", "Admin only"),
        ("GET", "/api/v1/health", "헬스체크", "없음"),
    ]
    for i, row in enumerate(api_data, start=23):
        for j, val in enumerate(row, start=1):
            c = ws2.cell(row=i, column=j, value=val)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                c.fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")

    ws2["A29"] = "※ 모든 API는 HTTPS 통신을 사용합니다."
    ws2["A29"].font = Font(size=9, italic=True, color="808080")

    # ── Sheet 3: Summary ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 40

    ws3["A1"] = "📊 요약 (Summary)"
    ws3["A1"].font = Font(bold=True, size=18, color="1F3864")

    ws3["A2"] = "전체 프로젝트 핵심 지표 요약"
    ws3["A2"].font = Font(size=11, italic=True)

    # Embedded image #3
    img3_buf = make_image("KPI Dashboard", width=400, height=180, bg=(142, 68, 173))
    xl_img3 = XLImage(img3_buf)
    xl_img3.width = 400
    xl_img3.height = 180
    ws3.add_image(xl_img3, "A4")

    kpi_headers = ["지표", "값"]
    for col, h in enumerate(kpi_headers, start=1):
        c = ws3.cell(row=13, column=col, value=h)
        header_style(c, "8E44AD")
        c.border = thin_border()

    kpi_data = [
        ("총 업무 수", "6개"),
        ("완료율", "33%"),
        ("진행중", "2개"),
        ("참여 인원", "5명"),
        ("예상 완료일", "2026-05-20"),
        ("예산 소진율", "42%"),
    ]
    for i, row in enumerate(kpi_data, start=14):
        for j, val in enumerate(row, start=1):
            c = ws3.cell(row=i, column=j, value=val)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                c.fill = PatternFill(start_color="F5EEF8", end_color="F5EEF8", fill_type="solid")

    wb.save("sample.xlsx")
    print("✅ sample.xlsx 생성 완료")
    print("   - Sheet 1: Project Overview  (텍스트 + 이미지 + 테이블)")
    print("   - Sheet 2: Technical Specs   (텍스트 + 병합셀 + 이미지 + 2개 테이블)")
    print("   - Sheet 3: Summary           (텍스트 + 이미지 + KPI 테이블)")


if __name__ == "__main__":
    create_sample()
